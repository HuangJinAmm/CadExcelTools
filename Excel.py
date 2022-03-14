import openpyxl
import xlrd
class Excel:

    def __init__(self,headPosition=(0,0),sheet=0,template=None):
        self.hp = headPosition
        self.sht = sheet 
        self.temp = template

    def readFromFile(self,fileName):
        if(fileName.lower().endswith("xlsx")):
            return self.readXlsx(fileName)
        else:
            return self.readXls(fileName)

    def readXlsx(self,fileName):
        retMap = {}
        retList = []
        wb = openpyxl.load_workbook(fileName,read_only=True)
        ws = wb.worksheets[self.sht]
        keys =list(ws.iter_rows(min_row = self.hp,max_row = self.hp+1))[0]
        for row in ws.iter_rows(min_row = self.hp+1,values_only=True):
            for key , cell in zip(keys,row):
                retMap[key.value]=cell
                retList.append(retMap)
        return retList

    def readXls(self,fileName):
        retMap = {}
        retList = []
        wb = xlrd.open_workbook(filename=fileName,on_demand=True)
        ws = wb.sheet_by_index(self.sht)
        keys =ws.row_values(self.hp)
        rows = ws.get_rows()
        for _ in range(self.hp+1):
            next(rows)
        while True:
            try:
                row = next(rows)
                for key ,cell in zip(keys,row):
                    retMap[key]=cell.value
                    retList.append(retMap)
            except StopIteration as identifier:
                break
        return retList

    def writeTofile(self,dataMap,root,fileName:str,filenamePost:str = None):
        
        if not bool(self.temp):
            wb = openpyxl.Workbook()
        else:
            wb = openpyxl.load_workbook(self.temp)

        ws = wb.active
        if bool(filenamePost):
            ws[filenamePost] = fileName[0:fileName.rindex(".")]
        if isinstance(dataMap[0],dict):
            for row in dataMap:
                for key,value in row.items():
                    ws[key] = value
        elif isinstance(dataMap[0],(list,tuple)):
            for i,row in enumerate(dataMap):
                for y,v in enumerate(row):
                    ws.cell(row=i+self.hp[0]+1,column=y+self.hp[1]+1,value=v)
        
        wb.save(root+fileName)

    
        


            


